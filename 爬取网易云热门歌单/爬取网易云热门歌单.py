from openpyxl import Workbook
from selenium import webdriver

url='https://music.163.com/#/discover/playlist/?cat=%E6%B5%81%E8%A1%8C'
driver=webdriver.Chrome()

wb = Workbook()  #创建一个工作簿的同时至少也新建一张工作表

# grab the active worksheet
ws = wb.active   #创建

# Data can be assigned directly to cells
ws['A1'] = '标题'   #写入单元格值
ws['B1'] = '播放数'
ws['C1'] = '链接'

# Rows can also be appended
#解析每一个网页，直到‘下一页’为空
while url !='javascript:void(0)':  #javascript:void(0)代表点击啥也发生的链接
    #用webdriver加载页面
    driver.get(url)
    #切换到内容的iframe
    driver.switch_to.frame("contentFrame")      #因为一个页面如果只有一个html，所有内容都要用表格来分是很难而且很丑，所以引
                                                # 进了frame可以将左侧树、右侧内容，顶端导航轻松分开。
                                                #么我们也有应对的方法就是要操作哪个元素先进入这个对应元素的frame里，一般frame有name或id属性。
    #定位歌单标签
    data=driver.find_element_by_id("m-pl-container").find_elements_by_tag_name("li")
    #id在HTML界面中是唯一的存在
    #任意打开一个页面，都会发现大量的<div>、<input>、<a>等tag
    #解析一页中的所有歌单
    for i in range(len(data)):
        #获取播放数
        nb=data[i].find_element_by_class_name("nb").text    #根据元素的class属性来定位
        if '万' in nb and int(nb.split("万")[0])>500:
            #获取播放数大于500万的歌单的封面
            msk=data[i].find_element_by_css_selector("a.msk")
            ##把封面上的标题和链接连同播放次数一起写到文件中
            #sys.stdout = io.TextIOWrapper(sys.stdout.buffer,encoding='gb18030')
            ws.append([msk.get_attribute('title'),nb,msk.get_attribute('href')])
         #  print([msk.get_attribute('title'),nb,msk.get_attribute('href')])
    #定位‘下一页’的url
    url=driver.find_element_by_css_selector("a.zbtn.znxt").get_attribute('href')

# Save the file
wb.save("sample.xlsx")